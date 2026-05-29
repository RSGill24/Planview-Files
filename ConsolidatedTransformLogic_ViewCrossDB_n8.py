"""
============================================================
ConsolidatedTransformLogic_ViewCrossDB.py
============================================================
Consolidated Transform Logic Pipeline — Planview 2026 Platinum

Steps:
  1. Read input Excel (3 sheets, row 3 header logic)
  2. Connect to SQL Server (Windows Auth)
  3. Create input_YYYYMMDD schema, load 3 tables
  4. Call dbo.Consolidated_Transform_Logic @run_ts
  5. Save output to PlanviewMigration_Transform_Output (versioned)
  6. Write Excel — colour coded + _Original columns + amber change highlighting
  7. Log to run_history.Pipeline_Runs_ConsolidatedTransform

Usage:
  py ConsolidatedTransformLogic_ViewCrossDB.py
  py ConsolidatedTransformLogic_ViewCrossDB.py --config "D:/other/config.json"

Requirements:
  pip install pandas pyodbc openpyxl
============================================================
"""

import sys
import io
import re
import json
import math
import argparse
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ── Globals ───────────────────────────────────────────────────────────────────
INPUT_FILE    = ""
INPUT_SHEETS  = {}
SQL_SERVER    = ""
SQL_DATABASE  = ""
OUTPUT_FOLDER = ""
EMAIL_ENABLED = False
EMAIL_TO      = []
EMAIL_CC      = []
EMAIL_SUBJECT = ""

STORED_PROC       = "dbo.Consolidated_Transform_Logic"
OUTPUT_SCHEMA     = "PlanviewMigration_Transform_Output"

# CrossDB view settings (populated by load_config)
VIEW_INITIATIVES  = "vw_Initiatives"
VIEW_EPICS        = "vw_Epics"
VIEW_SBA          = ""
VIEW_TASKS        = ""
VIEW_SCHEMA       = "dbo"
VIEW_DATABASE     = ""
COLUMN_RENAME     = {}    # populated by load_config()
MAX_CELL_LEN      = 32767
SEPARATOR         = "=" * 62
_logger           = None

# ── Colour constants ──────────────────────────────────────────────────────────
AMBER       = PatternFill("solid", start_color="FFE699")   # amber  — changed cells
YELLOW_HDR  = PatternFill("solid", start_color="FFFF00")   # yellow — generated col header
YELLOW_DATA = PatternFill("solid", start_color="FFFFE0")   # light yellow
ORANGE_HDR  = PatternFill("solid", start_color="FFC000")
ORANGE_DATA = PatternFill("solid", start_color="FFEEBA")
GREEN_HDR   = PatternFill("solid", start_color="92D050")
GREEN_DATA  = PatternFill("solid", start_color="E2EFDA")
BLUE_HDR    = PatternFill("solid", start_color="00B0F0")
BLUE_DATA   = PatternFill("solid", start_color="DDEEFF")
PURPLE_HDR  = PatternFill("solid", start_color="7030A0")
PURPLE_DATA = PatternFill("solid", start_color="EAD1F5")
BOLD_FONT   = Font(bold=True)


# ── Logging ───────────────────────────────────────────────────────────────────
def init_logger(ts):
    global _logger
    log_dir = Path(OUTPUT_FOLDER)
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"ConsolidatedTransform_Run_{ts}.log"
    _logger = logging.getLogger("ctl_crossdb_pipeline")
    _logger.setLevel(logging.DEBUG)
    _logger.handlers.clear()
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s  %(message)s",
                                      datefmt="%Y-%m-%d %H:%M:%S"))
    _logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(message)s"))
    _logger.addHandler(ch)
    return log_file


def log(msg, indent=0):
    line = "  " * indent + str(msg)
    _logger.info(line) if _logger else print(line)


def log_step(n, msg):
    line = f"\n[{n}] {msg}"
    _logger.info(line) if _logger else print(line)


# ── Config ────────────────────────────────────────────────────────────────────
def load_config(config_path=None):
    global VIEW_INITIATIVES, VIEW_EPICS, VIEW_SBA, VIEW_TASKS
    global VIEW_SCHEMA, VIEW_DATABASE
    global SQL_SERVER, SQL_DATABASE, OUTPUT_FOLDER
    global EMAIL_ENABLED, EMAIL_TO, EMAIL_CC, EMAIL_SUBJECT
    global COLUMN_RENAME

    if config_path is None:
        config_path = Path(__file__).parent / "Planview2026PlatinumConsolidated_ViewCrossDB_config.json"
    config_path = Path(config_path)
    if not config_path.exists():
        print(f"ERROR: Config not found: {config_path}"); sys.exit(1)
    try:
        with open(config_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: Bad JSON: {e}"); sys.exit(1)

    OUTPUT_FOLDER    = cfg["output_folder"]
    SQL_SERVER       = cfg.get("sql", {}).get("server",        "")
    SQL_DATABASE     = cfg.get("sql", {}).get("database",      "")
    VIEW_DATABASE    = cfg.get("sql", {}).get("view_database", "")
    VIEW_SCHEMA      = cfg.get("view_schema",      "dbo")
    VIEW_INITIATIVES = cfg.get("view_initiatives", "vw_Initiatives")
    VIEW_EPICS       = cfg.get("view_epics",       "vw_Epics")
    VIEW_SBA         = cfg.get("view_sba",         "")
    VIEW_TASKS       = cfg.get("view_tasks",       "")
    EMAIL_ENABLED    = cfg.get("email", {}).get("enabled", False)
    EMAIL_TO         = cfg.get("email", {}).get("to",      [])
    EMAIL_CC         = cfg.get("email", {}).get("cc",      [])
    EMAIL_SUBJECT    = cfg.get("email", {}).get("subject", "")
    COLUMN_RENAME    = cfg.get("column_rename", {})

    log(f"  Config        : {config_path}", 1)
    log(f"  View database : {VIEW_DATABASE}.{VIEW_SCHEMA}", 1)
    log(f"  Views         : {VIEW_INITIATIVES}, {VIEW_EPICS}", 1)
    log(f"  SQL Server    : {SQL_SERVER} / {SQL_DATABASE}", 1)
    log(f"  Output folder : {OUTPUT_FOLDER}", 1)


# ── Column name helpers ───────────────────────────────────────────────────────
def _sql_col(name):
    """
    Sanitise a column name for use inside SQL [...] brackets.
    The following characters ARE valid inside [...] in SQL Server
    and must NOT be stripped:
      '?'  — e.g. 'Is this confidential?'
      '('  — e.g. 'L1 Net Recurring Benefits ($, annualized)-P&L/Hard'
      ')'  — e.g. same
      '$'  — e.g. same
      ','  — e.g. same
      '-'  — e.g. same
      '/'  — e.g. same
    Only ']' breaks bracket-delimited identifiers and must be replaced.
    '[' inside a name is also replaced to avoid confusion.
    """
    clean = (str(name)
             .replace('\u2018', "'").replace('\u2019', "'")   # curly single quotes
             .replace('\u201c', '"').replace('\u201d', '"')   # curly double quotes
             .replace(']', ')')    # ] is the only char that breaks [...] names
             .replace('[', '(')    # [ inside a name replaced for clarity
             .replace('\\', '_')
             .strip())
    return clean[:120]


def _sql_type(series):
    """Infer SQL Server column type from pandas Series dtype."""
    dtype = series.dtype
    if pd.api.types.is_integer_dtype(dtype):   return 'bigint'
    if pd.api.types.is_float_dtype(dtype):     return 'decimal(20,2)'
    if pd.api.types.is_datetime64_any_dtype(dtype): return 'datetime'
    if pd.api.types.is_bool_dtype(dtype):      return 'bit'
    return 'nvarchar(MAX)'


def safe_val(v):
    """
    Convert a Python value to a SQL-safe value.
    - None / NaN / Inf  → None  (→ SQL NULL)
    - Empty string      → None  (→ SQL NULL)
    - String            → stripped, newlines/pipes cleaned
    - Numeric / date    → pass through as-is
    """
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    try:
        if not isinstance(v, (str, bool)) and pd.isnull(v):
            return None
    except Exception:
        pass
    if isinstance(v, str):
        s = (v.strip()
             .replace("\n", " ").replace("\r", " ")
             .replace("|",  " ").replace('"',  "'"))
        return s if s else None
    return v


def _sanitize_df(df):
    """
    Strip openpyxl-illegal characters and truncate to Excel's 32,767 char limit.
    Applied to every DataFrame before writing to Excel.
    """
    def clean(v):
        if v is None:
            return v
        s = str(v)
        s = ILLEGAL_CHARACTERS_RE.sub('', s)
        return s[:MAX_CELL_LEN]
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].apply(clean)
    return df


# ── Step 1: Read from SQL views (cross-database) ─────────────────────────────
def read_from_view(conn, view_name, label):
    """
    Read all rows from a SQL view using three-part name:
    [view_database].[view_schema].[view_name]
    Falls back to single-part name if VIEW_DATABASE not set.
    Returns DataFrame. Skips if view_name is blank.
    """
    if not view_name or not view_name.strip():
        log(f"  [{label}] — view not configured, skipping", 1)
        return pd.DataFrame()

    full_name = (f"[{VIEW_DATABASE}].[{VIEW_SCHEMA}].[{view_name}]"
                 if VIEW_DATABASE else f"[{VIEW_SCHEMA}].[{view_name}]")
    try:
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {full_name}")
        cols = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        df = pd.DataFrame.from_records([tuple(r) for r in rows], columns=cols)
        log(f"  {label:<15} {full_name} — {len(df):,} rows | {len(df.columns)} cols", 1)
        return df
    except Exception as e:
        log(f"  ERROR reading view {full_name}: {e}", 1)
        sys.exit(1)


def read_input(conn):
    log_step("1/7", "Reading data from SQL views...")
    df_init = read_from_view(conn, VIEW_INITIATIVES, "Initiatives")
    df_epic = read_from_view(conn, VIEW_EPICS,       "Epics")
    df_sba  = read_from_view(conn, VIEW_SBA,         "Value_Bundles")
    df_task = read_from_view(conn, VIEW_TASKS,       "Tasks")

    # Apply column renames from config — maps view col names to SP expected names
    # Update config file only if view column names change — no code change needed
    if not df_init.empty:
        df_init.rename(columns=COLUMN_RENAME.get("Initiatives", {}), inplace=True)
        renamed = COLUMN_RENAME.get("Initiatives", {})
        if renamed:
            log(f"  Initiatives — renamed {len(renamed)} columns per config", 1)
    if not df_epic.empty:
        df_epic.rename(columns=COLUMN_RENAME.get("Epics", {}), inplace=True)
        renamed = COLUMN_RENAME.get("Epics", {})
        if renamed:
            log(f"  Epics — renamed {len(renamed)} columns per config", 1)
    if not df_sba.empty:
        df_sba.rename(columns=COLUMN_RENAME.get("Value_Bundles", {}), inplace=True)
        renamed = COLUMN_RENAME.get("Value_Bundles", {})
        if renamed:
            log(f"  Value_Bundles — renamed {len(renamed)} columns per config", 1)

    # Drop pipeline metadata columns if view was built from a prior output table
    for df in [df_init, df_epic, df_sba, df_task]:
        for drop_col in ['Run_ID', 'Load_Timestamp', 'Save_Timestamp']:
            if drop_col in df.columns:
                df.drop(columns=[drop_col], inplace=True)

    frames = {}
    if not df_init.empty: frames["Initiatives"]   = df_init
    if not df_epic.empty: frames["Epics"]         = df_epic
    if not df_sba.empty:  frames["Value_Bundles"] = df_sba
    # Tasks not loaded to input schema for CTL pipeline
    return frames


# ── Step 2: Connect SQL ───────────────────────────────────────────────────────
def connect_sql():
    log_step("2/7", "Connecting to SQL Server...")
    try:
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
            f"Trusted_Connection=yes;Connection Timeout=30;"
        )
        conn.autocommit = True
        log(f"  Connected : {SQL_SERVER} / {SQL_DATABASE}", 1)
        return conn
    except Exception as e:
        log(f"  ERROR connecting: {e}", 1); sys.exit(1)


# ── Step 3: Load input schema ─────────────────────────────────────────────────
def load_to_sql(conn, df, schema, table, ts):
    """Load a DataFrame into SQL Server input schema table."""
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema}')
            EXEC('CREATE SCHEMA [{schema}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema}].[{table}]') IS NOT NULL
            DROP TABLE [{schema}].[{table}]
    """)

    # Sanitise column names and handle duplicates
    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols

    # All input columns as NVARCHAR(MAX) — SP handles types internally
    col_defs = ", ".join([f"[{c}] NVARCHAR(MAX)" for c in df.columns])
    cursor.execute(f"CREATE TABLE [{schema}].[{table}] ({col_defs})")

    # Verify column counts match
    cursor.execute(f"SELECT COUNT(*) FROM sys.columns "
                   f"WHERE object_id = OBJECT_ID('{schema}.{table}')")
    sql_col_count = cursor.fetchone()[0]
    if len(df.columns) != sql_col_count:
        log(f"  ERROR: Column count mismatch for {table}: "
            f"df={len(df.columns)} sql={sql_col_count}", 1)
        sys.exit(1)

    # Insert rows — fast_executemany OFF to avoid 510-char truncation
    rows = [tuple(safe_val(v) for v in row)
            for row in df.itertuples(index=False, name=None)]
    placeholders = ", ".join(["?"] * sql_col_count)
    ins = f"INSERT INTO [{schema}].[{table}] VALUES ({placeholders})"
    conn.autocommit = False
    try:
        cursor.fast_executemany = False
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema}].[{table}] — {len(rows):,} rows loaded", 1)


def dedup_input_table(conn, schema, table, key_col):
    """Remove duplicate rows in SQL input table on key_col."""
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT COUNT(*) - COUNT(DISTINCT [{key_col}])
        FROM [{schema}].[{table}]
        WHERE [{key_col}] IS NOT NULL
          AND LTRIM(RTRIM([{key_col}])) <> ''
    """)
    dup_count = cursor.fetchone()[0]
    if dup_count > 0:
        cursor.execute(f"""
            WITH CTE AS (
                SELECT ROW_NUMBER() OVER (
                    PARTITION BY [{key_col}]
                    ORDER BY (SELECT NULL)
                ) AS rn
                FROM [{schema}].[{table}]
                WHERE LTRIM(RTRIM(ISNULL([{key_col}], ''))) <> ''
            )
            DELETE FROM CTE WHERE rn > 1
        """)
        conn.commit()
        log(f"  Dedup [{table}]: removed {dup_count:,} duplicate rows "
            f"on [{key_col}]", 1)
    else:
        log(f"  Dedup [{table}]: no duplicates found", 1)


def create_input_schema(conn, run_ts, frames):
    log_step("3/7", "Loading input data to SQL Server...")
    schema = f"input_{run_ts}"
    dedup_keys = {
        "Initiatives":   "Strategy Seq ID",
        "Epics":         "Sequence ID",
        "Value_Bundles": "Sequence ID",
    }
    for table_name, df in frames.items():
        load_to_sql(conn, df, schema, table_name, run_ts)
        if table_name in dedup_keys:
            dedup_input_table(conn, schema, table_name, dedup_keys[table_name])
    log(f"  Input schema : [{schema}]", 1)
    return schema


# ── Step 4: Call stored procedure ─────────────────────────────────────────────
def call_stored_procedure(conn, run_ts):
    log_step("4/7", f"Calling {STORED_PROC}...")
    cursor = conn.cursor()
    conn.autocommit = False
    try:
        cursor.execute(f"EXEC {STORED_PROC} @run_ts = ?", run_ts)
    except Exception as e:
        log(f"  ERROR executing stored procedure: {e}", 1)
        log(f"  Make sure Consolidated_Transform_Logic.sql has been run "
            f"in SSMS first (CREATE OR ALTER PROCEDURE).", 1)
        conn.rollback()
        conn.autocommit = True
        sys.exit(1)

    # The SP runs SET NOCOUNT ON but still produces intermediate result sets
    # from ALTER TABLE / UPDATE statements in some driver versions.
    # Advance past any non-data result sets (cursor.description is None)
    # until we reach the first real SELECT result set.
    results = {}
    table_order = ['Initiatives', 'Epics', 'Value_Bundles']

    for table_name in table_order:
        # Skip intermediate empty result sets until we find one with columns
        while cursor.description is None:
            if not cursor.nextset():
                break

        if cursor.description:
            cols = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            df = pd.DataFrame.from_records([tuple(r) for r in rows], columns=cols)
            results[table_name] = df
            log(f"  {table_name:<15} → {len(df):,} rows | {len(df.columns)} cols", 1)
        else:
            log(f"  WARNING: No result set found for {table_name}", 1)
            results[table_name] = pd.DataFrame()

        # Advance to next result set for next table
        cursor.nextset()

    conn.commit()
    conn.autocommit = True
    return results


# ── Step 5: Save to output schema ─────────────────────────────────────────────
def get_next_version(conn):
    """Find next available version number in PlanviewMigration_Transform_Output."""
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{OUTPUT_SCHEMA}')
            EXEC('CREATE SCHEMA [{OUTPUT_SCHEMA}]')
    """)
    cursor.execute(f"""
        SELECT MAX(CAST(REPLACE(t.name, 'Initiatives_v', '') AS int))
        FROM sys.tables t
        JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE s.name = '{OUTPUT_SCHEMA}'
          AND t.name LIKE 'Initiatives_v%'
          AND ISNUMERIC(REPLACE(t.name, 'Initiatives_v', '')) = 1
    """)
    row = cursor.fetchone()
    max_v = row[0] if row and row[0] is not None else 0
    return max_v + 1



def align_output_types(conn, version):
    """
    After saving output as NVARCHAR(MAX), read view column metadata from
    sys.columns and ALTER each output table column to its correct SQL type.
    Only alters columns that exist in BOTH the view and the output table.
    New computed columns (Index_ID, Flow Type etc.) stay as NVARCHAR(MAX).
    """
    log("  Aligning output table column types to match view metadata...", 1)

    # Map of output table name → source view name
    view_map = {
        f"Initiatives_v{version}":   VIEW_INITIATIVES,
        f"Epics_v{version}":         VIEW_EPICS,
        f"Value_Bundles_v{version}": VIEW_SBA,
    }

    # SQL Server type map: system_type_id + max_length/precision/scale → SQL type string
    def build_type_str(type_name, max_length, precision, scale):
        type_name = type_name.lower()
        if type_name in ('nvarchar', 'nchar'):
            length = 'MAX' if max_length == -1 else str(max_length // 2)
            return f"{type_name}({length})"
        if type_name in ('varchar', 'char'):
            length = 'MAX' if max_length == -1 else str(max_length)
            return f"{type_name}({length})"
        if type_name in ('decimal', 'numeric'):
            return f"{type_name}({precision},{scale})"
        if type_name in ('float', 'real'):
            return type_name
        if type_name in ('int', 'bigint', 'smallint', 'tinyint',
                         'bit', 'datetime', 'datetime2', 'date',
                         'time', 'uniqueidentifier', 'money',
                         'smallmoney', 'image', 'text', 'ntext'):
            return type_name
        return 'nvarchar(MAX)'  # safe fallback for unknown types

    cursor = conn.cursor()

    for out_table, view_name in view_map.items():
        if not view_name or not view_name.strip():
            continue

        full_out = f"[{OUTPUT_SCHEMA}].[{out_table}]"
        if not cursor.tables(table=out_table, schema=OUTPUT_SCHEMA).fetchone():
            log(f"  {out_table} — not found, skipping type alignment", 1)
            continue

        # Get view column types from sys.columns
        view_db = VIEW_DATABASE if VIEW_DATABASE else SQL_DATABASE
        try:
            cursor.execute(f"""
                SELECT c.name, t.name as type_name,
                       c.max_length, c.precision, c.scale
                FROM [{view_db}].sys.columns c
                JOIN [{view_db}].sys.types t ON c.user_type_id = t.user_type_id
                JOIN [{view_db}].sys.objects o ON c.object_id = o.object_id
                JOIN [{view_db}].sys.schemas s ON o.schema_id = s.schema_id
                WHERE s.name = ? AND o.name = ?
            """, VIEW_SCHEMA, view_name)
            view_cols = {row[0]: build_type_str(row[1], row[2], row[3], row[4])
                         for row in cursor.fetchall()}
        except Exception as e:
            log(f"  WARNING: Could not read view metadata for {view_name}: {e}", 1)
            continue

        # Get output table columns
        cursor.execute(f"""
            SELECT c.name FROM sys.columns c
            JOIN sys.objects o ON c.object_id = o.object_id
            JOIN sys.schemas s ON o.schema_id = s.schema_id
            WHERE s.name = ? AND o.name = ?
        """, OUTPUT_SCHEMA, out_table)
        out_cols = {row[0] for row in cursor.fetchall()}

        # Apply renames from config — so we match the SP column names not view names
        rename_map = {}
        for table_key in ['Initiatives', 'Epics', 'Value_Bundles']:
            rename_map.update(COLUMN_RENAME.get(table_key, {}))
        # Build reverse map: SP name → view name (to look up view type)
        sp_to_view = {v: k for k, v in rename_map.items()}

        altered = 0
        skipped = []
        for col in out_cols:
            # Find the corresponding view column name
            view_col_name = sp_to_view.get(col, col)
            if view_col_name not in view_cols:
                continue  # Computed/new column — leave as NVARCHAR(MAX)

            sql_type = view_cols[view_col_name]
            if sql_type.lower() in ('nvarchar(max)', 'nvarchar(4000)'):
                continue  # Already correct, no need to alter

            try:
                cursor.execute(f"""
                    ALTER TABLE {full_out}
                    ALTER COLUMN [{col}] {sql_type} NULL
                """)
                altered += 1
            except Exception as e:
                skipped.append(f"{col} ({e})")

        conn.commit()
        log(f"  {out_table} — {altered} columns typed | "
            f"{len(skipped)} skipped", 1)
        if skipped:
            log(f"    Skipped: {skipped[:5]}", 1)


def save_to_output_schema(conn, results, run_ts):
    log_step("5/7", f"Saving output to [{OUTPUT_SCHEMA}]...")
    version = get_next_version(conn)
    v = f"v{version}"
    log(f"  Output version: {v}", 1)

    cursor = conn.cursor()
    for table_name, df in results.items():
        tbl = f"{table_name}_{v}"
        full = f"[{OUTPUT_SCHEMA}].[{tbl}]"
        cursor.execute(f"""
            IF OBJECT_ID('{OUTPUT_SCHEMA}.{tbl}') IS NOT NULL
                DROP TABLE {full}
        """)

        safe_cols = [_sql_col(c) for c in df.columns]
        seen = {}
        final_cols = []
        for c in safe_cols:
            if c in seen:
                seen[c] += 1
                c = f"{c}_{seen[c]}"
            else:
                seen[c] = 0
            final_cols.append(c)
        df_out = df.copy()
        df_out.columns = final_cols

        # All output columns as NVARCHAR(MAX) — SP result columns
        # come back as mixed/object dtype through pyodbc; using NVARCHAR
        # avoids all float/decimal type mismatch errors on insert.
        col_defs = ", ".join([f"[{c}] NVARCHAR(MAX)"
                              for c in df_out.columns])
        cursor.execute(f"""
            CREATE TABLE {full} (
                [Run_ID]         nvarchar(50) DEFAULT '{run_ts}',
                [Save_Timestamp] datetime     DEFAULT GETDATE(),
                {col_defs}
            )
        """)

        col_names    = ", ".join([f"[{c}]" for c in df_out.columns])
        placeholders = ", ".join(["?" for _ in df_out.columns])
        ins = (f"INSERT INTO {full} ([Run_ID],[Save_Timestamp],{col_names}) "
               f"VALUES ('{run_ts}',GETDATE(),{placeholders})")

        # Convert all columns to object/string dtype first —
        # this prevents pyodbc from sending float NaN as a typed float
        # parameter which SQL Server rejects for NVARCHAR columns.
        df_c = df_out.copy()
        df_c = df_c.where(df_c.notna(), None)  # NaN/NaT → None
        for col in df_c.columns:
            df_c[col] = df_c[col].apply(safe_val)
        rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]

        conn.autocommit = False
        try:
            for i in range(0, len(rows), 500):
                cursor.executemany(ins, rows[i:i+500])
            conn.commit()
        except Exception as e:
            conn.rollback(); raise
        finally:
            conn.autocommit = True
        log(f"  {full} — {len(rows):,} rows saved", 1)

    # Step 2: Align output column types to match source view metadata
    align_output_types(conn, version)

    return f"{OUTPUT_SCHEMA} ({v})"


# ── Step 6: Write Excel ───────────────────────────────────────────────────────
def add_original_cols(df_output, df_raw, mapped_cols, join_key):
    """
    Insert <col>_Original immediately after each mapped column.
    Compares output value to raw input value — NULL shown as blank not 'nan'.
    '?' column names are handled via fallback (raw col name + '?').
    """
    df = df_output.copy()
    if join_key not in df_raw.columns:
        return df

    dup_count = df_raw[join_key].duplicated().sum()
    if dup_count > 0:
        log(f"  WARNING: {dup_count:,} duplicate {join_key} values in raw data", 2)

    df_raw_dedup = df_raw.drop_duplicates(subset=[join_key], keep='first')
    raw_indexed  = df_raw_dedup.set_index(join_key)

    for col in mapped_cols:
        if col not in df.columns:
            continue
        # Try exact name then name+'?' (for cols where ? was stripped by _sql_col)
        raw_col = col if col in df_raw.columns else (
            col + '?' if col + '?' in df_raw.columns else None)
        if raw_col is None:
            continue

        orig_col = col + '_Original'
        col_idx  = df.columns.get_loc(col)
        orig_vals = (df[join_key]
                     .map(raw_indexed[raw_col])
                     .fillna('')
                     .astype(str)
                     .str.strip()
                     .replace('nan', ''))
        df.insert(col_idx + 1, orig_col, orig_vals.values)
    return df


def get_changed_positions(df_with_orig, mapped_cols):
    """Return set of (row, col) 1-based Excel positions where value changed."""
    changed = set()
    for col in mapped_cols:
        orig_col = col + '_Original'
        if col not in df_with_orig.columns or orig_col not in df_with_orig.columns:
            continue
        col_i  = df_with_orig.columns.get_loc(col) + 1
        orig_i = df_with_orig.columns.get_loc(orig_col) + 1
        for row_i, (nv, ov) in enumerate(
            zip(df_with_orig[col], df_with_orig[orig_col]), start=2
        ):
            nv_s = str(nv).strip() if nv is not None else ''
            ov_s = str(ov).strip() if ov is not None else ''
            if nv_s != ov_s:
                changed.add((row_i, col_i))
                changed.add((row_i, orig_i))
    return changed


def write_excel(results, frames, run_ts):
    log_step("6/7", "Writing Excel output...")

    out_folder = Path(OUTPUT_FOLDER)
    out_folder.mkdir(parents=True, exist_ok=True)
    out_file = out_folder / f"ConsolidatedTransform_Output_{run_ts}.xlsx"

    # Fully-highlighted generated columns (all rows highlighted, no comparison)
    GENERATED_COLS = {
        "Index_ID":                                (YELLOW_HDR, YELLOW_DATA, 18),
        "Flow Type - Step 1":                      (ORANGE_HDR, ORANGE_DATA, 28),
        "Estimated Annualized Value Range_Step 1": (GREEN_HDR,  GREEN_DATA,  32),
        "Estimated Annualized Value Range_Step 2": (GREEN_HDR,  GREEN_DATA,  32),
        "Estimated Annualized Value Range_Step 3": (GREEN_HDR,  GREEN_DATA,  32),
        "Estimated Value Range_Consolidated":      (BLUE_HDR,   BLUE_DATA,   32),
        "Flow Type":                               (PURPLE_HDR, PURPLE_DATA, 28),
        "Stage_New":                               (YELLOW_HDR, YELLOW_DATA, 20),
        "Work Type_New":                           (YELLOW_HDR, YELLOW_DATA, 24),
        "Does this require Limited Visibility?":   (YELLOW_HDR, YELLOW_DATA, 32),
    }

    # Mapped columns — amber highlighting on changed cells + _Original column
    MAPPED_COLS = {
        "Initiatives": [
            "Home Portfolio",
            "Impacted Portfolios",
            "Demand Domain or Portfolio",
            "Lifecycle Status",
            "Demand SubType",
            "Stage_New",
        ],
        "Epics": [
            "Home Domain/Portfolio",
            "Home Portfolio",
            "Impacted Portfolios",
            "Demand Domain or Portfolio",
            "Work Status",
        ],
        "Value_Bundles": [],
    }

    JOIN_KEYS = {
        "Initiatives":   "Strategy Seq ID",
        "Epics":         "Sequence ID",
        "Value_Bundles": "Sequence ID",
    }

    # Log row counts before writing
    for table_name, df_out in results.items():
        log(f"  {table_name:<15} — {len(df_out):,} rows to write", 1)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        sheets_written = 0
        for table_name, df_out in results.items():
            df_raw    = frames.get(table_name, pd.DataFrame())
            join_key  = JOIN_KEYS.get(table_name, "")
            m_cols    = MAPPED_COLS.get(table_name, [])

            # Write placeholder if DataFrame is empty
            if df_out.empty:
                pd.DataFrame({"Note": [f"No records returned for {table_name}"]}).to_excel(
                    writer, sheet_name=table_name, index=False)
                log(f"  {table_name:<15} — empty, written as placeholder sheet", 1)
                sheets_written += 1
                continue

            # Add _Original columns
            if not df_raw.empty and m_cols and join_key:
                df_final = add_original_cols(df_out, df_raw, m_cols, join_key)
                changed  = get_changed_positions(df_final, m_cols)
            else:
                df_final = df_out.copy()
                changed  = set()

            # Sanitise before writing
            df_final = _sanitize_df(df_final)
            df_final.to_excel(writer, sheet_name=table_name, index=False)
            sheets_written += 1

            ws   = writer.sheets[table_name]
            cols = list(df_final.columns)

            # Apply amber highlight on changed cells
            for (ri, ci) in changed:
                ws.cell(row=ri, column=ci).fill = AMBER

            # Apply full-column highlights for generated columns
            for col_name, (hdr_fill, dat_fill, col_width) in GENERATED_COLS.items():
                if col_name not in cols:
                    continue
                col_num = cols.index(col_name) + 1
                ws.cell(row=1, column=col_num).fill = hdr_fill
                ws.cell(row=1, column=col_num).font = BOLD_FONT
                for row_num in range(2, ws.max_row + 1):
                    ws.cell(row=row_num, column=col_num).fill = dat_fill
                ws.column_dimensions[
                    ws.cell(row=1, column=col_num).column_letter].width = col_width

            # Auto-fit all other columns
            for col_cells in ws.columns:
                if ws.column_dimensions[col_cells[0].column_letter].width in (0, None):
                    mx = max((len(str(c.value)) for c in col_cells if c.value),
                             default=10)
                    ws.column_dimensions[
                        col_cells[0].column_letter].width = min(mx + 4, 60)

            log(f"  {table_name:<15} — {len(df_final):,} rows, "
                f"{len(changed)//2:,} cells changed", 1)

    with open(out_file, 'wb') as f:
        f.write(buf.getvalue())

    log(f"  Output file  : {out_file}", 1)
    return out_file


# ── Step 7: Log run ───────────────────────────────────────────────────────────
def log_run(conn, run_ts, frames, results, output_schema,
            out_file, elapsed, status):
    log_step("7/7", "Logging run to run_history...")
    cursor = conn.cursor()

    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='run_history')
            EXEC('CREATE SCHEMA [run_history]')
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_ConsolidatedTransform','U') IS NULL
        CREATE TABLE run_history.Pipeline_Runs_ConsolidatedTransform (
            Run_ID              NVARCHAR(50)  PRIMARY KEY,
            Run_Timestamp       DATETIME      DEFAULT GETDATE(),
            Pipeline_Name       NVARCHAR(200),
            Input_File          NVARCHAR(500),
            Input_Schema        NVARCHAR(200),
            Output_Schema       NVARCHAR(200),
            Initiatives_In      INT,
            Epics_In            INT,
            ValueBundles_In     INT,
            Initiatives_Out     INT,
            Epics_Out           INT,
            ValueBundles_Out    INT,
            Output_File_Path    NVARCHAR(500),
            Runtime_Seconds     DECIMAL(10,1),
            Run_Status          NVARCHAR(50)
        )
    """)

    # Column guard — safe for re-runs against older table versions
    for col, col_type in [
        ("Initiatives_In",   "INT"),
        ("Epics_In",         "INT"),
        ("ValueBundles_In",  "INT"),
        ("Initiatives_Out",  "INT"),
        ("Epics_Out",        "INT"),
        ("ValueBundles_Out", "INT"),
        ("Output_Schema",    "NVARCHAR(200)"),
        ("Runtime_Seconds",  "DECIMAL(10,1)"),
        ("Run_Status",       "NVARCHAR(50)"),
        ("Output_File_Path", "NVARCHAR(500)"),
    ]:
        cursor.execute(f"""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA='run_history'
                  AND TABLE_NAME='Pipeline_Runs_ConsolidatedTransform'
                  AND COLUMN_NAME='{col}'
            )
            ALTER TABLE run_history.Pipeline_Runs_ConsolidatedTransform
            ADD [{col}] {col_type}
        """)
    conn.commit()

    cursor.execute("""
        INSERT INTO run_history.Pipeline_Runs_ConsolidatedTransform (
            Run_ID, Pipeline_Name, Input_File,
            Input_Schema, Output_Schema,
            Initiatives_In, Epics_In, ValueBundles_In,
            Initiatives_Out, Epics_Out, ValueBundles_Out,
            Output_File_Path, Runtime_Seconds, Run_Status
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        run_ts,
        'Consolidated Transform Logic Pipeline',
        INPUT_FILE,
        f'input_{run_ts}',
        output_schema,
        len(frames.get('Initiatives',   [])),
        len(frames.get('Epics',         [])),
        len(frames.get('Value_Bundles', [])),
        len(results.get('Initiatives',   [])),
        len(results.get('Epics',         [])),
        len(results.get('Value_Bundles', [])),
        str(out_file),
        round(elapsed, 1),
        status
    ))
    conn.commit()
    log(f"  Run logged : Run_ID={run_ts}  Status={status}", 1)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Consolidated Transform Logic Pipeline")
    parser.add_argument("--config", default=None,
                        help="Path to config JSON file")
    args  = parser.parse_args()
    start = datetime.now()
    run_ts = start.strftime("%Y%m%d_%H%M%S")

    print(SEPARATOR)
    print("  Consolidated Transform Logic Pipeline")
    print(f"  Run timestamp : {run_ts}")
    print(SEPARATOR)

    load_config(args.config)
    log_file = init_logger(run_ts)

    frames        = {}
    results       = {}
    out_file      = None
    output_schema = ""
    conn          = None

    try:
        conn          = connect_sql()
        frames        = read_input(conn)
        input_schema  = create_input_schema(conn, run_ts, frames)
        results       = call_stored_procedure(conn, run_ts)
        output_schema = save_to_output_schema(conn, results, run_ts)
        out_file      = write_excel(results, frames, run_ts)

        elapsed = (datetime.now() - start).total_seconds()
        log_run(conn, run_ts, frames, results, output_schema,
                out_file, elapsed, "SUCCESS")
        conn.close()

        log(f"\n{SEPARATOR}")
        log(f"  PIPELINE COMPLETE")
        log(f"  Input schema  : input_{run_ts}")
        log(f"  Output schema : {output_schema}")
        log(f"  Excel output  : {out_file}")
        log(f"  Log file      : {log_file}")
        log(f"  Runtime       : {round(elapsed, 1)}s")
        log(SEPARATOR)

        log(f"\n  {'Table':<20} {'Input':>8}  {'Output':>8}")
        log(f"  {'-'*40}")
        for t in ['Initiatives', 'Epics', 'Value_Bundles']:
            log(f"  {t:<20} {len(frames.get(t,[])):>8,}  "
                f"{len(results.get(t,[])):>8,}")

    except Exception as e:
        elapsed = (datetime.now() - start).total_seconds()
        log(f"\n  PIPELINE FAILED: {e}")
        if conn:
            try:
                log_run(conn, run_ts, frames, results, output_schema,
                        out_file, elapsed, f"FAILED: {str(e)[:200]}")
                conn.close()
            except Exception:
                pass
        sys.exit(1)


if __name__ == "__main__":
    main()
