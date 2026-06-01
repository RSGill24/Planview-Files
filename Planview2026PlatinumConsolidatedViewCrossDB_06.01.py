"""
Planview2026PlatinumConsolidatedViewCrossDB.py
===============================================
"""

import pandas as pd
import pyodbc
import json
import argparse
import sys
import io
import re
import logging
from pathlib import Path
from datetime import datetime

# ── Config globals ─────────────────────────────────────────────
VIEW_INITIATIVES  = "vw_Initiatives"
VIEW_EPICS        = "vw_Epics"
VIEW_SBA          = "vw_SBA"
VIEW_TASKS        = "vw_Tasks"
VIEW_SCHEMA       = "dbo"
VIEW_DATABASE     = ""        # source DB where views live
OUTPUT_FOLDER     = ""
SQL_SERVER        = ""
SQL_DATABASE      = ""        # working DB for input/output schemas
NRB_FIELD         = "L1 Net Recurring Benefits ($, annualized)-P&L/Hard"
NRB_THRESHOLD_M   = 10
_logger           = None
SEPARATOR         = "=" * 60


# ── Logging ───────────────────────────────────────────────────
def init_logger(ts, out_folder):
    global _logger
    Path(out_folder).mkdir(parents=True, exist_ok=True)
    log_file = Path(out_folder) / f"prod_pipeline_run_{ts}.log"
    _logger = logging.getLogger("prod_pipeline")
    _logger.setLevel(logging.DEBUG)
    _logger.handlers.clear()
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s  %(message)s",
                                       datefmt="%Y-%m-%d %H:%M:%S"))
    _logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(message)s"))
    _logger.addHandler(ch)
    return log_file

def log(msg, indent=0):
    line = "  " * indent + msg
    _logger.info(line) if _logger else print(line)

def log_step(n, msg):
    _logger.info(f"\n[{n}] {msg}") if _logger else print(f"\n[{n}] {msg}")


# ── Config loader ─────────────────────────────────────────────
def load_config(config_path=None):
    global VIEW_INITIATIVES, VIEW_EPICS, VIEW_SBA, VIEW_TASKS
    global VIEW_SCHEMA, VIEW_DATABASE
    global OUTPUT_FOLDER, SQL_SERVER, SQL_DATABASE, NRB_FIELD, NRB_THRESHOLD_M

    if config_path is None:
        config_path = Path(__file__).parent / "Planview2026PlatinumConsolidated_ViewCrossDB_config.json"
    config_path = Path(config_path)
    if not config_path.exists():
        print(f"ERROR: Config not found: {config_path}"); sys.exit(1)
    try:
        with open(config_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: Bad JSON: {e}"); sys.exit(1)

    OUTPUT_FOLDER     = cfg["output_folder"]
    SQL_SERVER        = cfg.get("sql", {}).get("server", "")
    SQL_DATABASE      = cfg.get("sql", {}).get("database", "")
    VIEW_DATABASE     = cfg.get("sql", {}).get("view_database", "")
    VIEW_SCHEMA       = cfg.get("view_schema", "dbo")
    VIEW_INITIATIVES  = cfg.get("view_initiatives", "vw_Initiatives")
    VIEW_EPICS        = cfg.get("view_epics",       "vw_Epics")
    VIEW_SBA          = cfg.get("view_sba",         "vw_SBA")
    VIEW_TASKS        = cfg.get("view_tasks",       "vw_Tasks")
    NRB_FIELD         = cfg.get("nrb_field",
                                "L1 Net Recurring Benefits ($, annualized)-P&L/Hard")
    NRB_THRESHOLD_M   = cfg.get("nrb_threshold_m", 10)


# ── Read from SQL view (cross-database) ──────────────────────
def read_from_view(conn, view_name):
    """
    Reads all rows from a SQL view using a three-part name:
        [view_database].[view_schema].[view_name]
    Falls back to single-part name if view_database is not set.
    Returns (df, type_map) where type_map is {col_name: sql_type_string}.
    """
    if not view_name or not view_name.strip():
        log(f"  [skipped — view name not configured]", 1)
        return pd.DataFrame(), {}
    if VIEW_DATABASE:
        full_name = f"[{VIEW_DATABASE}].[{VIEW_SCHEMA}].[{view_name}]"
    else:
        full_name = f"[{view_name}]"
    try:
        cursor = conn.cursor()

        # ── Fetch column type metadata from sys.columns in view database ──────
        type_map = {}
        try:
            if VIEW_DATABASE:
                meta_sql = f"""
                    SELECT c.name, t.name, c.precision, c.scale
                    FROM [{VIEW_DATABASE}].sys.columns c
                    JOIN [{VIEW_DATABASE}].sys.types t ON c.user_type_id = t.user_type_id
                    JOIN [{VIEW_DATABASE}].sys.objects o ON c.object_id = o.object_id
                    JOIN [{VIEW_DATABASE}].sys.schemas s ON o.schema_id = s.schema_id
                    WHERE o.name = ? AND s.name = ?
                    ORDER BY c.column_id
                """
                cursor.execute(meta_sql, (view_name, VIEW_SCHEMA))
            else:
                meta_sql = f"""
                    SELECT c.name, t.name, c.precision, c.scale
                    FROM sys.columns c
                    JOIN sys.types t ON c.user_type_id = t.user_type_id
                    JOIN sys.objects o ON c.object_id = o.object_id
                    JOIN sys.schemas s ON o.schema_id = s.schema_id
                    WHERE o.name = ? AND s.name = ?
                    ORDER BY c.column_id
                """
                cursor.execute(meta_sql, (view_name, VIEW_SCHEMA))
            for row in cursor.fetchall():
                col_name, type_name, prec, scale = row
                if type_name in ('decimal', 'numeric'):
                    type_map[col_name] = f'decimal({prec},{scale})'
                elif type_name in ('datetime', 'datetime2', 'date', 'smalldatetime'):
                    type_map[col_name] = 'datetime'
                elif type_name in ('int', 'bigint', 'smallint', 'tinyint'):
                    type_map[col_name] = 'bigint'
                elif type_name == 'bit':
                    type_map[col_name] = 'bit'
                elif type_name == 'float':
                    type_map[col_name] = 'float'
                else:
                    type_map[col_name] = 'nvarchar(MAX)'
        except Exception as meta_e:
            log(f"  Warning: could not fetch type metadata for [{view_name}]: {meta_e}", 1)
            type_map = {}

        # ── Fetch data ────────────────────────────────────────────────────────
        cursor.execute(f"SELECT * FROM {full_name}")
        cols = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        df = pd.DataFrame.from_records([tuple(r) for r in rows], columns=cols)
        log(f"  {full_name} — {len(df):,} rows | {len(df.columns)} cols", 1)
        return df, type_map
    except Exception as e:
        log(f"ERROR reading view {full_name}: {e}", 1)
        sys.exit(1)


# ── SQL: Connect ──────────────────────────────────────────────
def connect_sql():
    log_step("2/5", "Connecting to SQL Server...")
    try:
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
            f"Trusted_Connection=yes;Connection Timeout=30;"
        )
        conn.autocommit = True
        log(f"Server   : {SQL_SERVER} / {SQL_DATABASE}", 1)
        log("Status   : Connected", 1)
        return conn
    except pyodbc.Error as e:
        log(f"ERROR: Could not connect to SQL Server: {e}", 1)
        sys.exit(1)


# ── SQL column name sanitiser ─────────────────────────────────
def _sql_col(name):
    # Strips characters that are illegal in SQL column names even inside brackets.
    # Note: '?' IS valid inside [brackets] in SQL Server — do NOT strip it.
    clean = (str(name)
             .replace(']', '')
             .replace('[', '')
             .replace(',', '_')
             .replace('/', '_')
             .replace('\\', '_')
             .replace(':', '_')
             .replace('(', '')
             .replace(')', '')
             .strip())
    return clean[:120]


def _sql_type(series):
    """Map a pandas Series dtype to the appropriate SQL Server data type."""
    import pandas as pd
    dtype = series.dtype
    if pd.api.types.is_integer_dtype(dtype):
        return 'bigint'
    elif pd.api.types.is_float_dtype(dtype):
        return 'decimal(20,2)'
    elif pd.api.types.is_datetime64_any_dtype(dtype):
        return 'datetime'
    elif pd.api.types.is_bool_dtype(dtype):
        return 'bit'
    else:
        return 'nvarchar(MAX)'


# ── SQL: Load single dataframe into a table ───────────────────
def load_to_sql(conn, df, schema_name, table_name, ts, type_map=None):
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema_name}')
            EXEC('CREATE SCHEMA [{schema_name}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema_name}].[{table_name}]') IS NOT NULL
            DROP TABLE [{schema_name}].[{table_name}]
    """)

    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols

    # Drop pipeline metadata cols if view was built on a previously saved output table
    df = df.drop(columns=[c for c in ['Run_ID', 'Load_Timestamp', 'Save_Timestamp']
                           if c in df.columns])

    col_defs = ", ".join([
        f"[{c}] {type_map.get(c, 'nvarchar(MAX)') if type_map else _sql_type(df[c])}"
        for c in df.columns
    ])
    cursor.execute(f"""
        CREATE TABLE [{schema_name}].[{table_name}] (
            [Run_ID]         nvarchar(50)  DEFAULT '{ts}',
            [Load_Timestamp] datetime      DEFAULT GETDATE(),
            {col_defs}
        )
    """)
    col_names    = ", ".join([f"[{c}]" for c in final_cols])
    placeholders = ", ".join(["?" for _ in final_cols])
    ins = (
        f"INSERT INTO [{schema_name}].[{table_name}] "
        f"([Run_ID],[Load_Timestamp],{col_names}) "
        f"VALUES ('{ts}',GETDATE(),{placeholders})"
    )
    df_c = df.copy()

    def safe_val(v, col):
        """Return None for nulls (→ SQL NULL), typed value for numerics/dates, str for text."""
        import math, pandas as pd
        if v is None:
            return None
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if pd.isnull(v) if not isinstance(v, (str, bool)) else False:
            return None
        # For string columns clean the value
        if isinstance(v, str):
            s = (v.strip()
                 .replace("\n", " ").replace("\r", " ")
                 .replace("|", " ").replace('"', "'"))
            return s if s else None
        return v  # numeric/date — pass through as-is

    for col in df_c.columns:
        df_c[col] = df_c[col].apply(lambda v: safe_val(v, col))

    rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]
    conn.autocommit = False
    try:
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema_name}].[{table_name}] — {len(rows):,} rows loaded", 1)


# ── SQL: Deduplicate input table on a key column ──────────────
def dedup_input_table(conn, schema, table, key_col, label):
    """
    Removes duplicate rows from an input table keeping the first occurrence
    per key_col. Uses ROW_NUMBER() to identify and delete duplicates.
    Logs how many rows were removed.
    """
    cursor = conn.cursor()
    full_table = f"[{schema}].[{table}]"

    # Check if key column exists
    cursor.execute(f"""
        SELECT COUNT(*) FROM sys.columns
        WHERE object_id = OBJECT_ID('{schema}.{table}')
        AND name = '{key_col}'
    """)
    if cursor.fetchone()[0] == 0:
        log(f"  Dedup skipped — [{key_col}] not found in {label}", 1)
        return

    # Count duplicates before
    cursor.execute(f"""
        SELECT COUNT(*) - COUNT(DISTINCT [{key_col}])
        FROM {full_table}
        WHERE [{key_col}] IS NOT NULL AND LTRIM(RTRIM([{key_col}])) <> ''
    """)
    dup_count = cursor.fetchone()[0]

    if dup_count > 0:
        # Delete keeping first occurrence per key using ROW_NUMBER
        cursor.execute(f"""
            WITH CTE AS (
                SELECT ROW_NUMBER() OVER (
                    PARTITION BY [{key_col}]
                    ORDER BY (SELECT NULL)
                ) AS rn
                FROM {full_table}
                WHERE LTRIM(RTRIM(ISNULL([{key_col}], ''))) <> ''
            )
            DELETE FROM CTE WHERE rn > 1
        """)
        conn.commit()
        log(f"  Dedup — {label}: removed {dup_count:,} duplicate rows "
            f"(kept first occurrence per [{key_col}])", 1)
    else:
        log(f"  Dedup — {label}: no duplicates found", 1)


# ── SQL: Load raw input into input schema ─────────────────────
def create_input_schema_and_load(conn, df_inits, df_epics, df_sba, df_tasks, ts,
                                  type_map_init=None, type_map_epic=None,
                                  type_map_sba=None, type_map_task=None):
    log_step("2a/5", "Loading view data to SQL Server input schema...")
    schema   = f"input_{ts}"
    stem     = "Planview_View"
    tbl_init = f"{stem}_Initiatives"
    tbl_epic = f"{stem}_Epics"
    tbl_sba  = f"{stem}_SBA"
    tbl_task = f"{stem}_Tasks"
    load_to_sql(conn, df_inits, schema, tbl_init, ts, type_map_init)
    load_to_sql(conn, df_epics, schema, tbl_epic, ts, type_map_epic)
    if not df_sba.empty:
        load_to_sql(conn, df_sba, schema, tbl_sba, ts, type_map_sba)
    else:
        log(f"  SBA view not configured — skipping load", 1)
    if not df_tasks.empty:
        load_to_sql(conn, df_tasks, schema, tbl_task, ts, type_map_task)
    else:
        # SP requires the Tasks table to exist — create an empty placeholder
        log(f"  Tasks view not configured — creating empty placeholder table", 1)
        cursor = conn.cursor()
        cursor.execute(f"""
            IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema}')
                EXEC('CREATE SCHEMA [{schema}]')
        """)
        cursor.execute(f"""
            IF OBJECT_ID('[{schema}].[{tbl_task}]') IS NOT NULL
                DROP TABLE [{schema}].[{tbl_task}]
        """)
        cursor.execute(f"""
            CREATE TABLE [{schema}].[{tbl_task}] (
                [_placeholder] nvarchar(10) NULL
            )
        """)
        conn.commit()

    # ── Deduplicate input tables in SQL before SP runs ────────────────────────
    # Ensures SP only processes one row per record — fixes mapping inconsistencies
    # caused by duplicate rows in the view.
    log("Deduplicating input tables in SQL...", 1)
    dedup_input_table(conn, schema, tbl_init, 'Strategy Seq ID', 'Initiatives')
    dedup_input_table(conn, schema, tbl_epic, 'Sequence ID',     'Epics')
    dedup_input_table(conn, schema, tbl_task, 'Sequence ID',     'Tasks')

    log(f"Input schema : [{schema}]", 1)
    return schema, stem


# ── STEPS 2a–4: Call stored procedure ────────────────────────
def run_transform_sp(conn, ts, input_schema, stem):
    """
    Runs the mapping stored procedure and returns mapped dataframes.

    Returns:
        df_inits     — mapped Initiatives dataframe
        df_epics     — mapped Epics dataframe
        changes_in   — dict {col: cnt}
        changes_ep   — dict {col: cnt}
    """
    log_step("3/5", "Planview2026PlatinumConsolidatedUpdated...")

    nrb_field_sql = _sql_col(NRB_FIELD)

    cursor = conn.cursor()
    cursor.execute(
        "EXEC dbo.Planview2026PlatinumConsolidatedUpdated "
        "    @RunID=?, @InputSchema=?, @Stem=?, @NRB_Field=?, @NRB_Threshold_M=?",
        ts, input_schema, stem, nrb_field_sql, float(NRB_THRESHOLD_M)
    )

    # ── Result set 1: mapped Initiatives ─────────────────────────────────────
    cols_in = [col[0] for col in cursor.description]
    rows_in = cursor.fetchall()
    df_inits = pd.DataFrame.from_records(
        [tuple(r) for r in rows_in], columns=cols_in
    )
    df_inits = df_inits.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_inits.columns]
    )

    # ── Result set 2: mapped Epics ────────────────────────────────────────────
    cursor.nextset()
    cols_ep = [col[0] for col in cursor.description]
    rows_ep = cursor.fetchall()
    df_epics = pd.DataFrame.from_records(
        [tuple(r) for r in rows_ep], columns=cols_ep
    )
    df_epics = df_epics.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp'] if c in df_epics.columns]
    )

    # ── Result set 3: scalar counts (all 0) ──────────────────────────────────
    cursor.nextset()
    cursor.fetchone()  # consume RS3

    # ── Result set 4: changes_in ──────────────────────────────────────────────
    cursor.nextset()
    changes_in = {}
    for r in cursor.fetchall():
        if r.cnt and r.cnt > 0:
            changes_in[r.col] = r.cnt

    # ── Result set 5: changes_ep ──────────────────────────────────────────────
    cursor.nextset()
    changes_ep = {}
    for r in cursor.fetchall():
        if r.cnt and r.cnt > 0:
            changes_ep[r.col] = r.cnt

    # ── Result set 6: mapped Tasks ────────────────────────────────────────────
    cursor.nextset()
    cols_tk = [col[0] for col in cursor.description]
    rows_tk = cursor.fetchall()
    df_tasks = pd.DataFrame.from_records(
        [tuple(r) for r in rows_tk], columns=cols_tk
    )
    df_tasks = df_tasks.drop(
        columns=[c for c in ['Run_ID', 'Load_Timestamp', '_placeholder'] if c in df_tasks.columns]
    )

    log(f"Step 2a — Value transformations:", 1)
    if changes_in:
        for col, cnt in changes_in.items():
            log(f"Initiatives — {col}: {cnt:,} values remapped", 2)
    if changes_ep:
        for col, cnt in changes_ep.items():
            log(f"Epics       — {col}: {cnt:,} values remapped", 2)
    if not changes_in and not changes_ep:
        log("No values remapped", 2)

    log(f"Mapping complete:", 1)
    log(f"Initiatives : {len(df_inits):,} rows | {len(df_inits.columns)} cols", 2)
    log(f"Epics       : {len(df_epics):,} rows | {len(df_epics.columns)} cols", 2)
    log(f"Tasks       : {len(df_tasks):,} rows | {len(df_tasks.columns)} cols", 2)

    return df_inits, df_epics, df_tasks, changes_in, changes_ep


# ── SQL: Save single output table ─────────────────────────────
def save_output_to_sql(conn, df, schema_name, table_name, ts):
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema_name}')
            EXEC('CREATE SCHEMA [{schema_name}]')
    """)
    cursor.execute(f"""
        IF OBJECT_ID('[{schema_name}].[{table_name}]') IS NOT NULL
            DROP TABLE [{schema_name}].[{table_name}]
    """)
    if df.empty:
        log(f"  [{schema_name}].[{table_name}] — 0 rows (skipped)", 1)
        return

    safe_cols = [_sql_col(c) for c in df.columns]
    seen = {}
    final_cols = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        final_cols.append(c)
    df = df.copy()
    df.columns = final_cols

    # Drop pipeline metadata cols if view was built on a previously saved output table
    df = df.drop(columns=[c for c in ['Run_ID', 'Load_Timestamp', 'Save_Timestamp']
                           if c in df.columns])

    col_defs = ", ".join([f"[{c}] {_sql_type(df[c])}" for c in df.columns])
    cursor.execute(f"""
        CREATE TABLE [{schema_name}].[{table_name}] (
            [Run_ID]         nvarchar(50) DEFAULT '{ts}',
            [Save_Timestamp] datetime     DEFAULT GETDATE(),
            {col_defs}
        )
    """)
    col_names    = ", ".join([f"[{c}]" for c in df.columns])
    placeholders = ", ".join(["?" for _ in df.columns])
    ins = (
        f"INSERT INTO [{schema_name}].[{table_name}] "
        f"([Run_ID],[Save_Timestamp],{col_names}) "
        f"VALUES ('{ts}',GETDATE(),{placeholders})"
    )
    import math, pandas as pd

    def safe_val(v, col):
        if v is None:
            return None
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        try:
            if not isinstance(v, (str, bool)) and pd.isnull(v):
                return None
        except Exception:
            pass
        if isinstance(v, str):
            s = (v.strip()
                 .replace("\n", " ").replace("\r", " ")
                 .replace("|", " ").replace('"', "'"))
            return s if s else None
        return v

    df_c = df.copy()
    for col in df_c.columns:
        df_c[col] = df_c[col].apply(lambda v: safe_val(v, col))
    rows = [tuple(r) for r in df_c.itertuples(index=False, name=None)]
    conn.autocommit = False
    try:
        for i in range(0, len(rows), 500):
            cursor.executemany(ins, rows[i:i+500])
        conn.commit()
    except Exception as e:
        conn.rollback(); raise
    finally:
        conn.autocommit = True
    log(f"  [{schema_name}].[{table_name}] — {len(df):,} rows saved", 1)


# ── SQL: Save all output sets to output schema ────────────────
def get_next_version(conn, schema):
    """Find the next available version number for output tables."""
    cursor = conn.cursor()
    cursor.execute(f"""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='{schema}')
            EXEC('CREATE SCHEMA [{schema}]')
    """)
    cursor.execute(f"""
        SELECT MAX(CAST(REPLACE(name, 'Initiatives_v', '') AS int))
        FROM sys.tables t
        JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE s.name = '{schema}'
        AND t.name LIKE 'Initiatives_v%'
        AND ISNUMERIC(REPLACE(t.name, 'Initiatives_v', '')) = 1
    """)
    row = cursor.fetchone()
    max_v = row[0] if row and row[0] is not None else 0
    return max_v + 1


def create_output_schema_and_save(conn, df_inits, df_epics, df_sba, df_tasks, stem, ts):
    log_step("5a/5", "Saving mapped output to SQL Server...")
    schema  = "PlanviewMigration_Output"
    version = get_next_version(conn, schema)
    v       = f"v{version}"
    log(f"Output version: {v}", 1)
    save_output_to_sql(conn, df_inits, schema, f"Initiatives_{v}", ts)
    save_output_to_sql(conn, df_epics, schema, f"Epics_{v}",       ts)
    save_output_to_sql(conn, df_sba,   schema, f"SBA_{v}",         ts)
    save_output_to_sql(conn, df_tasks, schema, f"Tasks_{v}",       ts)
    log(f"Output schema: [{schema}] — tables versioned as {v}", 1)
    return f"{schema} ({v})"

# ── Helper: Add _Original columns next to transformed columns ─
def add_original_cols(df_output, df_raw, mapped_cols, join_key):
    """
    Inserts <col>_Original immediately after each transformed column.
    Uses join_key to match output rows back to the correct input row.
    Handles duplicate join key values by keeping first occurrence.
    Handles cases where _sql_col() strips special chars (e.g. '?') from
    column names — tries the raw name and the raw name + '?' as fallback.
    NULL/None values from view data are shown as blank (not 'nan').
    """
    df = df_output.copy()

    if join_key not in df_raw.columns:
        return df

    # Warn if view is returning duplicate join key values
    dup_count = df_raw[join_key].duplicated().sum()
    if dup_count > 0:
        log(f"  WARNING: {dup_count:,} duplicate {join_key} values in raw data — "
            f"view may be returning multiple rows per record. "
            f"_Original values will use first occurrence.", 2)

    # Deduplicate on join_key — keep first occurrence to avoid InvalidIndexError
    df_raw_dedup = df_raw.drop_duplicates(subset=[join_key], keep='first')
    raw_indexed  = df_raw_dedup.set_index(join_key)

    for col in mapped_cols:
        if col not in df.columns:
            continue

        raw_col = None
        if col in df_raw.columns:
            raw_col = col
        elif col + '?' in df_raw.columns:
            raw_col = col + '?'

        if raw_col is None:
            continue

        orig_col = col + '_Original'
        col_idx  = df.columns.get_loc(col)

        orig_vals = (df[join_key]
            .map(raw_indexed[raw_col])
            .fillna('')
            .astype(str)
            .str.strip()
            .replace('nan', '')   # NULL from view comes through as 'nan' — show as blank
        )

        df.insert(col_idx + 1, orig_col, orig_vals.values)

    return df


def get_changed_positions(df_with_orig, mapped_cols):
    """
    Returns set of (row, col) 1-based Excel positions where value changed.
    Compares <col> vs <col>_Original for each mapped column.
    Row 1 = header so data starts at row 2.
    """
    changed = set()
    for col in mapped_cols:
        orig_col = col + '_Original'
        if col not in df_with_orig.columns or orig_col not in df_with_orig.columns:
            continue
        col_i  = df_with_orig.columns.get_loc(col) + 1
        orig_i = df_with_orig.columns.get_loc(orig_col) + 1
        for row_i, (nv, ov) in enumerate(
            zip(df_with_orig[col], df_with_orig[orig_col]), start=2
        ):
            nv_s = str(nv).strip() if nv is not None else ''
            ov_s = str(ov).strip() if ov is not None else ''
            if nv_s != ov_s:
                changed.add((row_i, col_i))
                changed.add((row_i, orig_i))
    return changed


# ── STEP 5b: Build Excel output ───────────────────────────────
def build_excel(df_inits, df_epics, df_sba, df_tasks, df_in_raw, df_ep_raw, df_tk_raw,
                ts, input_path, out_folder, changes_in, changes_ep):

    from openpyxl.styles import PatternFill
    AMBER = PatternFill("solid", start_color="FFE699", fgColor="FFE699")

    def _sanitize_df(df):
        """
        Strip all characters that openpyxl cannot write to Excel cells.
        Uses openpyxl's ILLEGAL_CHARACTERS_RE directly to strip illegal chars,
        then truncates to Excel's 32,767 character cell limit.
        """
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        MAX_CELL = 32767

        def clean(v):
            if v is None:
                return v
            s = str(v)
            # Strip illegal characters using openpyxl's own pattern
            s = ILLEGAL_CHARACTERS_RE.sub('', s)
            # Truncate to Excel cell limit
            return s[:MAX_CELL]

        df = df.copy()
        for col in df.columns:
            df[col] = df[col].apply(clean)
        return df

    def write_sheet(writer, df, sheet_name, df_raw=None, mapped_cols=None, join_key=None):
        if df.empty:
            pd.DataFrame({"Note": ["No records in this category"]}).to_excel(
                writer, sheet_name=sheet_name, index=False)
            return
        if df_raw is not None and mapped_cols and join_key:
            df_out  = add_original_cols(df, df_raw, mapped_cols, join_key)
            changed = get_changed_positions(df_out, mapped_cols)
        else:
            df_out  = df
            changed = set()
        # Sanitize AFTER add_original_cols so _Original columns are also cleaned
        df_out = _sanitize_df(df_out)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for (ri, ci) in changed:
            ws.cell(row=ri, column=ci).fill = AMBER
        for col_cells in ws.columns:
            mx = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(mx + 4, 60)

    output_file = Path(out_folder) / f"Planview_Prod_Migration_Output_{ts}.xlsx"
    Path(out_folder).mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        init_mapped   = ['Estimated Annualized Value Range',
                         'Work Type', 'Home Portfolio', 'Risk Subtype',
                         'Impacted Portfolios', 'Demand Domain or Portfolio',
                         'Does this require Limited Visibility?', 'Status']
        epic_mapped   = ['Work Type', 'Work Status', 'Level of Governance',
                         'Home Domain/Portfolio', 'Impacted Portfolios',
                         'Demand Domain or Portfolio']
        task_mapped   = ['Task or Milestone Type']
        init_join_key = 'Strategy Seq ID'
        epic_join_key = 'Sequence ID'
        task_join_key = 'Sequence ID'

        # Sanitize raw dataframes too — _Original columns come from these
        df_in_raw  = _sanitize_df(df_in_raw)
        df_ep_raw  = _sanitize_df(df_ep_raw)
        df_tk_raw  = _sanitize_df(df_tk_raw)

        write_sheet(writer, df_inits, "Initiatives",
                    df_in_raw, init_mapped, init_join_key)
        write_sheet(writer, df_epics, "Epics",
                    df_ep_raw, epic_mapped, epic_join_key)
        write_sheet(writer, df_tasks, "Tasks",
                    df_tk_raw, task_mapped, task_join_key)
        # SBA — passthrough, no mappings applied, no highlighting
        write_sheet(writer, df_sba, "SBA")

        summary = [
            ["RUN INFORMATION", ""],
            ["Run Timestamp", ts],
            ["Input File",    str(input_path)],
            ["", ""],
            ["VALUE MAPPINGS APPLIED (Initiatives)", ""],
        ]
        for col, cnt in changes_in.items():
            summary.append([f"  {col}", f"{cnt:,} values remapped"])
        if not changes_in:
            summary.append(["  None", ""])

        summary += [["", ""], ["VALUE MAPPINGS APPLIED (Epics)", ""]]
        for col, cnt in changes_ep.items():
            summary.append([f"  {col}", f"{cnt:,} values remapped"])
        if not changes_ep:
            summary.append(["  None", ""])

        summary += [
            ["", ""],
            ["INITIATIVES", f"{len(df_inits):,} total"],
            ["EPICS",       f"{len(df_epics):,} total"],
            ["TASKS",       f"{len(df_tasks):,} total"],
            ["SBA",         f"{len(df_sba):,} total (passthrough — no mappings applied)"],
        ]

        df_sum = pd.DataFrame(summary, columns=["Item", "Value"])
        df_sum.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]
        ws_sum.column_dimensions['A'].width = 50
        ws_sum.column_dimensions['B'].width = 60

    with open(output_file, 'wb') as f:
        f.write(buf.getvalue())

    return output_file


# ── SQL: Log run to run_history ───────────────────────────────
def log_run_history(conn, ts, input_path, input_schema, output_schema,
                    df_inits, df_epics, df_sba, df_tasks, out_path, elapsed, status):
    cursor = conn.cursor()
    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name='run_history')
            EXEC('CREATE SCHEMA [run_history]')
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_Prod') IS NOT NULL
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM sys.columns
                WHERE object_id=OBJECT_ID('run_history.Pipeline_Runs_Prod')
                AND name='Task_Count'
            )
            DROP TABLE run_history.Pipeline_Runs_Prod
        END
    """)
    cursor.execute("""
        IF OBJECT_ID('run_history.Pipeline_Runs_Prod') IS NULL
        CREATE TABLE run_history.Pipeline_Runs_Prod (
            Run_ID             nvarchar(50)  PRIMARY KEY,
            Run_Timestamp      datetime      DEFAULT GETDATE(),
            Pipeline_Name      nvarchar(200),
            Pipeline_Version   nvarchar(20),
            Input_File         nvarchar(500),
            Input_Schema       nvarchar(200),
            Output_Schema      nvarchar(200),
            Init_Count         int,
            Epic_Count         int,
            SBA_Count          int,
            Task_Count         int,
            Output_File_Path   nvarchar(500),
            Runtime_Seconds    decimal(10,1),
            Run_Status         nvarchar(50)
        )
    """)
    cursor.execute("""
        INSERT INTO run_history.Pipeline_Runs_Prod (
            Run_ID, Pipeline_Name, Pipeline_Version, Input_File,
            Input_Schema, Output_Schema,
            Init_Count, Epic_Count, SBA_Count, Task_Count,
            Output_File_Path, Runtime_Seconds, Run_Status
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        ts, 'Prod Data Pipeline', 'prod_v6', str(input_path),
        input_schema, output_schema,
        len(df_inits), len(df_epics), len(df_sba), len(df_tasks),
        str(out_path), elapsed, status
    ))
    conn.commit()
    log(f"Run logged : run_history.Pipeline_Runs_Prod — Run_ID: {ts}", 1)


# ── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Planview Prod Data Pipeline — View mode")
    parser.add_argument("--config", default=None)
    args = parser.parse_args()
    load_config(args.config)

    start = datetime.now()
    ts    = start.strftime("%Y%m%d_%H%M%S")
    log_file = init_logger(ts, OUTPUT_FOLDER)

    log(SEPARATOR)
    log("  Planview2026PlatinumConsolidatedViewCrossDB (Cross-DB View mode)")
    log(f"  Views    : {VIEW_DATABASE}.{VIEW_SCHEMA} on {SQL_SERVER}")
    log(f"  Working  : {SQL_DATABASE} on {SQL_SERVER}")
    log(f"  Output   : {OUTPUT_FOLDER}")
    log(f"  Started  : {start.strftime('%Y-%m-%d %H:%M:%S')}")
    log(SEPARATOR)

    # ── Step 1 — Connect to SQL and read views ────────────────
    log_step("1/5", "Connecting to SQL and reading views...")
    conn = connect_sql()

    df_in_raw,  type_map_init = read_from_view(conn, VIEW_INITIATIVES)
    df_ep_raw,  type_map_epic = read_from_view(conn, VIEW_EPICS)
    df_sba_raw, type_map_sba  = read_from_view(conn, VIEW_SBA)
    df_tk_raw,  type_map_task = read_from_view(conn, VIEW_TASKS)

    log(f"Initiatives : {len(df_in_raw):,} rows | {len(df_in_raw.columns)} cols", 1)
    log(f"Epics       : {len(df_ep_raw):,} rows | {len(df_ep_raw.columns)} cols", 1)
    log(f"Tasks       : {len(df_tk_raw):,} rows | {len(df_tk_raw.columns)} cols", 1)
    log(f"SBA         : {len(df_sba_raw):,} rows | {len(df_sba_raw.columns)} cols (passthrough)", 1)

    # ── Step 2 — Load view data to SQL input schema ───────────
    input_schema, stem = create_input_schema_and_load(
        conn, df_in_raw, df_ep_raw, df_sba_raw, df_tk_raw, ts,
        type_map_init, type_map_epic, type_map_sba, type_map_task)

    # ── Step 3 — Stored procedure (mappings) ──────────────────
    # SBA bypasses the SP entirely — passed through as-is
    (df_inits, df_epics, df_tasks,
     changes_in, changes_ep) = run_transform_sp(conn, ts, input_schema, stem)

    # ── Deduplicate on join keys — remove duplicate rows from view ────────────
    # View may return multiple rows per record causing inconsistent mappings.
    # Keep first occurrence which has correctly mapped values.
    def dedup(df, key, label):
        if key in df.columns:
            before = len(df)
            df = df.drop_duplicates(subset=[key], keep='first').reset_index(drop=True)
            removed = before - len(df)
            if removed > 0:
                log(f"  Deduplication — {label}: removed {removed:,} duplicate rows "
                    f"(keeping first occurrence per {key})", 1)
        return df

    df_inits   = dedup(df_inits,   'Strategy Seq ID', 'Initiatives')
    df_epics   = dedup(df_epics,   'Sequence ID',     'Epics')
    df_tasks   = dedup(df_tasks,   'Sequence ID',     'Tasks')
    df_in_raw  = dedup(df_in_raw,  'Strategy Seq ID', 'Initiatives raw')
    df_ep_raw  = dedup(df_ep_raw,  'Sequence ID',     'Epics raw')
    df_tk_raw  = dedup(df_tk_raw,  'Sequence ID',     'Tasks raw')

    # ── Console summary ───────────────────────────────────────
    log(f"\nINITIATIVES ({len(df_inits):,} total)", 1)
    log(f"EPICS       ({len(df_epics):,} total)", 1)
    log(f"TASKS       ({len(df_tasks):,} total)", 1)
    log(f"SBA         ({len(df_sba_raw):,} total — passthrough)", 1)

    # ── Step 5a — Save mapped output to SQL ──────────────────
    output_schema = create_output_schema_and_save(
        conn, df_inits, df_epics, df_sba_raw, df_tasks, stem, ts)

    # ── Step 5b — Log run ─────────────────────────────────────
    log_step("5b/5", "Logging run to run_history.Pipeline_Runs_Prod...")

    # ── Step 5c — Write Excel output ─────────────────────────
    log_step("5c/5", "Writing output Excel...")
    out_path = build_excel(
        df_inits, df_epics, df_sba_raw, df_tasks,
        df_in_raw, df_ep_raw, df_tk_raw,
        ts, Path(f"view_{ts}"), OUTPUT_FOLDER, changes_in, changes_ep)

    elapsed = round((datetime.now() - start).total_seconds(), 1)

    log(f"\n{SEPARATOR}")
    log("  PIPELINE COMPLETE (Cross-DB View mode)")
    log(f"  Output : {out_path}")
    log(f"  Log    : {log_file}")
    log(f"")
    log(f"  INITIATIVES : {len(df_inits):,} total")
    log(f"  EPICS       : {len(df_epics):,} total")
    log(f"  TASKS       : {len(df_tasks):,} total")
    log(f"  SBA         : {len(df_sba_raw):,} total (passthrough)")
    log(f"  Runtime     : {elapsed}s")

    log_run_history(conn, ts,
                    Path(f"views:{VIEW_DATABASE}.{VIEW_SCHEMA}:{VIEW_INITIATIVES},{VIEW_EPICS},{VIEW_TASKS},{VIEW_SBA}"),
                    input_schema, output_schema,
                    df_inits, df_epics, df_sba_raw, df_tasks,
                    out_path, elapsed, "Completed")

    conn.close()
    log("SQL      : Connection closed", 1)
    log(SEPARATOR)


if __name__ == "__main__":
    main()
